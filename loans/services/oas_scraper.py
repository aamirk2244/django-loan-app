"""
Django view: extract balance data from Askari Bank PDF statements
and save as Excel in data/oas-amount/

- Processes ALL PDFs in data/sample-statements/
- Multiple PDFs per account → uses the one with the latest statement period end date
- Fallback ordering: latest "To: DD-MON-YY" date, then latest filename (alphabetical desc)

Setup:
    pip install pdfplumber openpyxl

Add to urls.py:
    from .oas_extractor import process_oas_statements
    path("api/process-oas/", process_oas_statements, name="process_oas"),
"""

import re
from datetime import datetime
from pathlib import Path

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.conf import settings


STATEMENTS_DIR = Path("data/sample-statements")
OAS_AMOUNT_DIR = Path("data/oas-amount")

MONTH_MAP = {"JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
             "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_statement_end_date(text: str) -> datetime | None:
    """
    Extract the 'To: DD-MON-YY' date from the statement header.
    Returns a datetime or None if not found.
    """
    m = re.search(r"To\s*[:\-]?\s*(\d{2})[-/\s]([A-Z]{3})[-/\s](\d{2,4})", text, re.IGNORECASE)
    if not m:
        return None
    day, mon, yr = m.group(1), m.group(2).upper(), m.group(3)
    year = int(yr) + 2000 if len(yr) == 2 else int(yr)
    month = MONTH_MAP.get(mon)
    if not month:
        return None
    return datetime(year, month, int(day))


def extract_text_from_pdf(pdf_path: Path) -> str:
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text += t
    return text


def extract_account_number(text: str) -> str:
    # Line like: "038AKJL00000666  Term Finance - ..."
    m = re.search(r"^([\w]{10,25})\s+Term Finance", text, re.MULTILINE)
    if m:
        return m.group(1).strip()
    # Fallback: ACCOUNT NUMBER header
    m2 = re.search(r"ACCOUNT\s+NUMBER\s*\n?\s*([\w]+)", text)
    return m2.group(1).strip() if m2 else "UNKNOWN"


def extract_balances(text: str) -> dict:
    """
    Walk lines and pick the BALANCE column value (line immediately after each
    PRINCIPAL REPAYMENT row) for JAN, FEB, MAR.

    Business rules:
      - Multiple entries in a month → use the last one.
      - Last entry is 0 AND earlier non-zero entries exist → use last non-zero.
      - Month entirely absent → 0.0.
      - JAN absent → fall back to Opening Balance value.
    """
    lines = text.splitlines()
    month_balance: dict[str, list[float]] = {
        "JAN": [], "FEB": [], "MAR": []
    }

    for i, line in enumerate(lines):
        m = re.match(r"\d{2}-(JAN|FEB|MAR)-\d{2}\s+PRINCIPAL REPAYMENT", line.strip())
        if m:
            month = m.group(1)
            if i + 1 < len(lines):
                nxt = lines[i + 1].strip()
                bm = re.match(r"^(-?[\d,]+\.\d{2})$", nxt)
                if bm:
                    month_balance[month].append(float(bm.group(1).replace(",", "")))

    def pick(vals: list[float]) -> float | None:
        if not vals:
            return None
        last = vals[-1]
        if last == 0 and len(vals) > 1:
            non_zero = [v for v in vals if v != 0]
            return non_zero[-1] if non_zero else 0.0
        return last

    jan = pick(month_balance["JAN"])
    feb = pick(month_balance["FEB"]) or 0.0
    mar = pick(month_balance["MAR"]) or 0.0

    if jan is None:
        ob = re.search(r"Opening Balance\s*\*\*\s*(-?[\d,]+\.\d{2})", text)
        jan = float(ob.group(1).replace(",", "")) if ob else 0.0

    return {"jan": jan, "feb": feb, "mar": mar}


def extract_statement_data(pdf_path: Path) -> dict:
    """Parse a single PDF. Returns account_number, statement_end_date, balances."""
    text = extract_text_from_pdf(pdf_path)
    return {
        "account_number":   extract_account_number(text),
        "statement_end":    parse_statement_end_date(text),   # datetime | None
        "filename":         pdf_path.name,
        **extract_balances(text),
    }


def pick_latest_per_account(records: list[dict]) -> list[dict]:
    """
    For each account number, keep only the record with the latest statement.
    Priority: statement_end date (desc) → filename (desc, alphabetical).
    """
    grouped: dict[str, list[dict]] = {}
    for rec in records:
        grouped.setdefault(rec["account_number"], []).append(rec)

    result = []
    for acct, recs in grouped.items():
        def sort_key(r):
            # Use a very old date as fallback so None sorts last
            dt = r["statement_end"] or datetime(2000, 1, 1)
            return (dt, r["filename"])
        best = max(recs, key=sort_key)
        result.append(best)

    # Sort final output by account number for a clean Excel
    result.sort(key=lambda r: r["account_number"])
    return result


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def save_to_excel(records: list[dict], output_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OAS Amounts"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    data_font    = Font(name="Arial", size=10)
    even_fill    = PatternFill("solid", start_color="DCE6F1")
    odd_fill     = PatternFill("solid", start_color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="BFBFBF")
    cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fmt      = '#,##0.00;(#,##0.00);"-"'

    headers    = [
        "Account Number",
        "Statement End Date",
        "Source File",
        "Jan Balance (PKR)",
        "Feb Balance (PKR)",
        "Mar Balance (PKR)",
    ]
    col_widths = [24, 20, 36, 22, 22, 22]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = cell_border
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for r_idx, rec in enumerate(records, start=2):
        fill = even_fill if r_idx % 2 == 0 else odd_fill
        end_date = rec["statement_end"].strftime("%d-%b-%Y") if rec["statement_end"] else "N/A"
        row_vals = [
            rec["account_number"],
            end_date,
            rec["filename"],
            rec["jan"],
            rec["feb"],
            rec["mar"],
        ]
        for c_idx, val in enumerate(row_vals, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.fill   = fill
            c.border = cell_border
            c.font   = data_font
            if c_idx in (1, 2, 3):
                c.alignment = left_align
            else:
                c.alignment    = center_align
                c.number_format = num_fmt

    wb.save(output_path)

import re
import pandas as pd


def merge_oas_amounts_to_master(master_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    """
    Appends month balance columns from new_df onto master_df by matching
    'PMKJ Master Table_Loan_Acc' (master) to 'Account Number' (new_df).
    No existing master data is modified — only new columns are added.
    """
    MASTER_ID_COL = "PMKJ Master Table_Loan_Acc"
    NEW_ID_COL    = "Account Number"

    if MASTER_ID_COL not in master_df.columns:
        raise ValueError(f"Master file is missing identifier column: '{MASTER_ID_COL}'")
    if NEW_ID_COL not in new_df.columns:
        raise ValueError(f"OAS file is missing identifier column: '{NEW_ID_COL}'")

    # Dynamically detect month balance columns in new_df
    balance_pattern = re.compile(
        r"\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
        r"jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\b",
        re.IGNORECASE,
    )
    oas_cols = [col for col in new_df.columns if balance_pattern.search(col)]

    if not oas_cols:
        raise ValueError(f"OAS file has no recognisable month columns. Found: {list(new_df.columns)}")

    # Keep only account number + balance columns from new_df
    new_subset = (
        new_df[[NEW_ID_COL] + oas_cols]
        .drop_duplicates(subset=[NEW_ID_COL])
        .rename(columns={NEW_ID_COL: MASTER_ID_COL})
    )

    # Left join — master rows stay intact, balance columns appended where account matches
    result = master_df.merge(new_subset, on=MASTER_ID_COL, how="left")

    return result
